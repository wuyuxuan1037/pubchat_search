import logging
import re
import xmltodict
import pandas as pd
import csv
import requests
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure the logger for this module
logger = logging.getLogger(__name__)

# 🔄 Unpaywall Email 轮换管理器
class UnpaywallEmailManager:
    """
    Unpaywall API Email 轮换管理器
    
    Unpaywall API 限制：每个 email 每天 100,000 次请求
    通过多 email 轮换突破限制，支持高并发场景
    """
    _instance = None
    _emails = []
    _current_index = 0
    
    @classmethod
    def initialize(cls, emails_str: str = None):
        """
        初始化 email 列表
        
        Args:
            emails_str: 逗号分隔的 email 列表（从环境变量读取）
        """
        import os
        if emails_str is None:
            emails_str = os.getenv("UNPAYWALL_EMAILS", "research@pubchat.org")
        
        cls._emails = [e.strip() for e in emails_str.split(",") if e.strip()]
        
        if not cls._emails:
            cls._emails = ["research@pubchat.org"]
        
        logger.info(f"📧 Unpaywall Email Manager initialized with {len(cls._emails)} email(s)")
    
    @classmethod
    def get_next_email(cls) -> str:
        """
        获取下一个 email（轮换）
        
        Returns:
            用于 Unpaywall API 的 email
        """
        if not cls._emails:
            cls.initialize()
        
        email = cls._emails[cls._current_index]
        cls._current_index = (cls._current_index + 1) % len(cls._emails)
        return email

# Excel 非法字符正则表达式（控制字符，除了 Tab, LF, CR）
ILLEGAL_CHARACTERS_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]'
)


def clean_illegal_characters(value):
    """
    清理 Excel 不允许的非法字符

    Args:
        value: 任意值

    Returns:
        清理后的值（如果是字符串则移除非法字符，否则原样返回）
    """
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub('', value)
    return value

# 全局缓存：期刊参考数据
_journal_ref_cache: Dict[str, Dict[str, str]] = {}
_journal_ref_loaded: bool = False


# ============================================================
# 期刊过滤相关函数
# ============================================================

def _normalize_comparison_symbols(expr: str) -> str:
    """
    统一比较符号（中英文转换）
    """
    replacements = {
        '＞': '>',
        '＜': '<',
        '≥': '>=',
        '≤': '<=',
        '～': '-',  # 全角波浪线 → 连字符
        '~': '-',   # 半角波浪线 → 连字符
    }
    for old, new in replacements.items():
        expr = expr.replace(old, new)
    return expr.strip()


def parse_if_expression(expr: str) -> Dict[str, Any]:
    """
    解析影响因子表达式

    支持格式:
        "3~5" 或 "3-5" -> 范围 3 到 5
        ">4" -> 大于 4
        ">=5" -> 大于等于 5
        "<8" -> 小于 8
        "<=8" -> 小于等于 8
        "5" -> 等于 5

    Returns:
        {"min": float, "max": float, "min_inclusive": bool, "max_inclusive": bool}
    """
    expr = _normalize_comparison_symbols(expr)

    # 范围格式: 3~5 或 3-5（但要排除负数如 -1）
    range_match = re.match(r'^(\d+\.?\d*)\s*[~\-]\s*(\d+\.?\d*)$', expr)
    if range_match:
        return {
            "min": float(range_match.group(1)),
            "max": float(range_match.group(2)),
            "min_inclusive": True,
            "max_inclusive": True
        }

    # >=
    if expr.startswith('>='):
        return {"min": float(expr[2:]), "max": float('inf'), "min_inclusive": True, "max_inclusive": True}

    # <=
    if expr.startswith('<='):
        return {"min": 0, "max": float(expr[2:]), "min_inclusive": True, "max_inclusive": True}

    # >
    if expr.startswith('>'):
        return {"min": float(expr[1:]), "max": float('inf'), "min_inclusive": False, "max_inclusive": True}

    # <
    if expr.startswith('<'):
        return {"min": 0, "max": float(expr[1:]), "min_inclusive": True, "max_inclusive": False}

    # 纯数字（精确匹配）
    if re.match(r'^\d+\.?\d*$', expr):
        val = float(expr)
        return {"min": val, "max": val, "min_inclusive": True, "max_inclusive": True}

    raise ValueError(f"无法解析影响因子表达式: {expr}")


def parse_jcr_expression(expr: str) -> List[str]:
    """
    解析 JCR 分区表达式

    Q1 最好，Q4 最差

    支持格式:
        "Q1" -> ["Q1"]
        "Q1-Q3" -> ["Q1", "Q2", "Q3"]
        ">Q3" -> ["Q1", "Q2", "Q3"] (比 Q3 好，含 Q3)
        ">=Q2" -> ["Q1", "Q2"]
        "<Q2" -> ["Q3", "Q4"]

    Returns:
        List of JCR zones, e.g., ["Q1", "Q2"]
    """
    expr = _normalize_comparison_symbols(expr).upper()
    all_zones = ["Q1", "Q2", "Q3", "Q4"]

    # 范围格式: Q1-Q3
    range_match = re.match(r'^Q(\d)\s*-\s*Q(\d)$', expr)
    if range_match:
        start = int(range_match.group(1))
        end = int(range_match.group(2))
        return [f"Q{i}" for i in range(start, end + 1)]

    # >=Qn（Qn 及更好）
    if expr.startswith('>='):
        n = int(expr[3])
        return [f"Q{i}" for i in range(1, n + 1)]

    # >Qn（比 Qn 好，含 Qn，即 Q1 到 Qn）
    if expr.startswith('>'):
        n = int(expr[2])
        return [f"Q{i}" for i in range(1, n + 1)]

    # <=Qn（Qn 及更差）
    if expr.startswith('<='):
        n = int(expr[3])
        return [f"Q{i}" for i in range(n, 5)]

    # <Qn（比 Qn 差，不含 Qn）
    if expr.startswith('<'):
        n = int(expr[2])
        return [f"Q{i}" for i in range(n + 1, 5)]

    # 单个分区: Q1
    single_match = re.match(r'^Q(\d)$', expr)
    if single_match:
        return [expr]

    raise ValueError(f"无法解析 JCR 分区表达式: {expr}")


def parse_cas_expression(expr: str) -> List[str]:
    """
    解析中科院分区表达式

    1区 最好，4区 最差

    支持格式:
        "1" -> ["1"]
        "1-3" -> ["1", "2", "3"]
        ">3" -> ["1", "2", "3"] (比 3 区好，含 3 区)
        ">=2" -> ["1", "2"]
        "<2" -> ["3", "4"]

    Returns:
        List of CAS zones, e.g., ["1", "2"]
    """
    expr = _normalize_comparison_symbols(expr)

    # 范围格式: 1-3
    range_match = re.match(r'^(\d)\s*-\s*(\d)$', expr)
    if range_match:
        start = int(range_match.group(1))
        end = int(range_match.group(2))
        return [str(i) for i in range(start, end + 1)]

    # >=n（n区 及更好）
    if expr.startswith('>='):
        n = int(expr[2])
        return [str(i) for i in range(1, n + 1)]

    # >n（比 n区 好，含 n区）
    if expr.startswith('>'):
        n = int(expr[1])
        return [str(i) for i in range(1, n + 1)]

    # <=n（n区 及更差）
    if expr.startswith('<='):
        n = int(expr[2])
        return [str(i) for i in range(n, 5)]

    # <n（比 n区 差，不含 n区）
    if expr.startswith('<'):
        n = int(expr[1])
        return [str(i) for i in range(n + 1, 5)]

    # 单个分区: 1
    single_match = re.match(r'^(\d)$', expr)
    if single_match:
        return [expr]

    raise ValueError(f"无法解析中科院分区表达式: {expr}")


def parse_journal_filters(journal_filters: Dict[str, str], logger: logging.Logger) -> Dict[str, Any]:
    """
    解析所有期刊过滤条件

    Args:
        journal_filters: {"impact_factor": "≥5", "jcr_zone": "Q1-Q2", "cas_zone": "1-2"}

    Returns:
        {
            "if_condition": {"min": 5, "max": inf, ...},
            "jcr_zones": ["Q1", "Q2"],
            "cas_zones": ["1", "2"]
        }
    """
    parsed = {}

    if journal_filters.get("impact_factor"):
        try:
            parsed["if_condition"] = parse_if_expression(journal_filters["impact_factor"])
        except ValueError as e:
            logger.warning(f"⚠️ {e}")

    if journal_filters.get("jcr_zone"):
        try:
            parsed["jcr_zones"] = parse_jcr_expression(journal_filters["jcr_zone"])
        except ValueError as e:
            logger.warning(f"⚠️ {e}")

    if journal_filters.get("cas_zone"):
        try:
            parsed["cas_zones"] = parse_cas_expression(journal_filters["cas_zone"])
        except ValueError as e:
            logger.warning(f"⚠️ {e}")

    return parsed


def load_journal_reference_for_filter(reference_file: str, logger: logging.Logger) -> Dict[str, Dict[str, Any]]:
    """
    加载期刊参考数据用于过滤（带缓存）

    Returns:
        Dict keyed by normalized ISSN: {"if": float, "jcr": str, "cas": str}
    """
    global _journal_ref_cache, _journal_ref_loaded

    if _journal_ref_loaded:
        return _journal_ref_cache

    try:
        # 检测编码
        try:
            with open(reference_file, 'r', encoding='utf-8') as f:
                f.read()
            encoding = 'utf-8'
        except UnicodeDecodeError:
            encoding = 'gbk'

        with open(reference_file, 'r', encoding=encoding) as f:
            reader = csv.reader(f)
            next(reader)  # Skip header
            for row in reader:
                if len(row) < 11:
                    continue

                # 提取 IF、JCR、CAS
                try:
                    if_value = float(row[8].strip()) if row[8].strip() else 0.0
                except ValueError:
                    if_value = 0.0

                jcr_zone = row[7].strip().upper()  # Q1, Q2, etc.
                cas_zone = row[6].strip()  # 1, 2, etc.

                five_year_if = row[9].strip() if len(row) > 9 else ""
                ranking = row[10].strip() if len(row) > 10 else ""

                journal_info = {
                    "if": if_value,
                    "jcr": jcr_zone,
                    "cas": cas_zone,
                    "journal_name": row[0].strip(),
                    "five_year_if": five_year_if,
                    "ranking": ranking
                }

                # 用 ISSN 和 EISSN 作为 key
                print_issn = row[2].strip().replace('-', '')
                e_issn = row[3].strip().replace('-', '')

                if print_issn:
                    _journal_ref_cache[print_issn] = journal_info
                if e_issn:
                    _journal_ref_cache[e_issn] = journal_info

        _journal_ref_loaded = True
        logger.info(f"📚 Loaded {len(_journal_ref_cache)} journal entries for filtering.")

    except Exception as e:
        logger.error(f"❌ Failed to load journal reference for filtering: {e}")

    return _journal_ref_cache


def filter_articles_by_journal(
    articles: List[Dict[str, Any]],
    journal_filters: Dict[str, str],
    reference_file: str,
    logger: logging.Logger
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    """
    根据期刊条件过滤文章，并补充期刊信息（IF、分区等）。
    如果没有过滤条件，则仅补充信息，不进行过滤。

    Args:
        articles: 解析后的文章列表（包含 issn 字段）
        journal_filters: 原始过滤条件 {"impact_factor": "≥5", ...}
        reference_file: 期刊参考数据文件路径

    Returns:
        (filtered_articles, stats)
    """
    # 加载期刊参考数据 (Always load to enable enrichment)
    journal_ref = load_journal_reference_for_filter(reference_file, logger)

    # 解析过滤条件
    parsed_filters = None
    if journal_filters:
        parsed_filters = parse_journal_filters(journal_filters, logger)
    
    filtered = []
    stats = {"before": len(articles), "after": 0, "no_issn": 0, "not_found": 0, "filtered_out": 0}

    for article in articles:
        issn = article.get("issn", "")
        if not issn:
            # 如果没有 ISSN
            if parsed_filters: # 有过滤要求则排除
                stats["no_issn"] += 1
                continue
            else: # 无过滤要求则保留
                filtered.append(article)
                continue

        # 标准化 ISSN
        normalized_issn = issn.replace('-', '').strip()

        # 查找期刊信息
        journal_info = journal_ref.get(normalized_issn)
        
        passed = True
        
        if parsed_filters:
            # --- FILTERING MODE ---
            if not journal_info:
                stats["not_found"] += 1
                continue # 找不到期刊信息，且有过滤要求 -> 排除

            # 检查所有过滤条件
            # 检查影响因子
            if "if_condition" in parsed_filters:
                cond = parsed_filters["if_condition"]
                journal_if = journal_info["if"]

                if cond["min_inclusive"]:
                    if journal_if < cond["min"]: passed = False
                else:
                    if journal_if <= cond["min"]: passed = False

                if passed and cond["max"] != float('inf'):
                    if cond["max_inclusive"]:
                        if journal_if > cond["max"]: passed = False
                    else:
                        if journal_if >= cond["max"]: passed = False

            # 检查 JCR 分区
            if passed and "jcr_zones" in parsed_filters:
                journal_jcr = journal_info["jcr"]
                if journal_jcr not in parsed_filters["jcr_zones"]:
                    passed = False

            # 检查中科院分区
            if passed and "cas_zones" in parsed_filters:
                journal_cas = journal_info["cas"]
                if journal_cas not in parsed_filters["cas_zones"]:
                    passed = False
        else:
            # --- NO FILTER MODE ---
            # 仅记录统计，不排除
            if not journal_info:
                stats["not_found"] += 1
                # passed stays True

        if passed:
            # Enrich article if info found
            if journal_info:
                article.update({
                    "latest_if": journal_info["if"],
                    "jcr_zone": journal_info["jcr"],
                    "cas_zone": journal_info["cas"],
                    "five_year_if": journal_info["five_year_if"],
                    "ranking": journal_info["ranking"]
                })
            filtered.append(article)
        else:
            stats["filtered_out"] += 1

    stats["after"] = len(filtered)

    # 输出日志
    if parsed_filters:
        logger.info(f"📰 Journal filtering: {stats['before']} → {stats['after']} articles")
        logger.info(f"   • Passed: {stats['after']} | Filtered out: {stats['filtered_out']}")
        logger.info(f"   • No ISSN: {stats['no_issn']} | Not in reference: {stats['not_found']}")
    else:
        logger.info(f"📰 Journal enrichment: {stats['before']} articles processed (No filters applied)")
        logger.info(f"   • Enriched (Found in DB): {stats['before'] - stats['not_found'] - stats['no_issn']} | Not found: {stats['not_found']}")

    return filtered, stats


def get_journal_filter_display(journal_filters: Dict[str, str]) -> str:
    """
    获取期刊过滤条件的显示字符串

    Args:
        journal_filters: {"impact_factor": "≥5", "jcr_zone": "Q1-Q2", "cas_zone": "1-2"}

    Returns:
        "IF≥5, JCR: Q1-Q2, CAS: 1-2区"
    """
    parts = []

    if journal_filters.get("impact_factor"):
        parts.append(f"IF{journal_filters['impact_factor']}")

    if journal_filters.get("jcr_zone"):
        parts.append(f"JCR: {journal_filters['jcr_zone']}")

    if journal_filters.get("cas_zone"):
        parts.append(f"CAS: {journal_filters['cas_zone']}区")

    return ", ".join(parts) if parts else ""


# ============================================================
# XML 解析相关函数
# ============================================================

def parse_pubmed_xml(xml_data: str, logger: logging.Logger) -> List[Dict[str, Any]]:
    """
    Parses raw XML data from PubMed into a list of article dictionaries.
    """
    try:
        logger.info("⚙️ Parsing PubMed XML data...")
        data_dict = xmltodict.parse(xml_data)
        articles = data_dict.get("PubmedArticleSet", {}).get("PubmedArticle", [])
        if isinstance(articles, dict):
            articles = [articles]
        logger.info(f"✅ Successfully parsed {len(articles)} articles from XML.")
        return articles
    except Exception as e:
        logger.error(f"❌ Failed to parse PubMed XML: {e}", exc_info=True)
        return []


# ============================================================
# Europe PMC JSON 解析相关函数
# ============================================================

def parse_europepmc_json(json_data: Dict[str, Any], logger: logging.Logger) -> Tuple[List[Dict[str, Any]], List[str], List[str]]:
    """
    解析 Europe PMC API 返回的 JSON 数据
    
    Args:
        json_data: Europe PMC API 返回的 JSON 响应
        logger: 日志记录器
    
    Returns:
        Tuple[List[Dict], List[str], List[str]]:
            - 有摘要的文献元数据列表（可直接用于 AI 筛选）
            - 无摘要但近90天的 PMID 列表（需要 PubMed 补充）
            - 无摘要且超90天的 PMID 列表（直接丢弃）
    """
    import html
    from ..clients.EuropePMCClient import EuropePMCClient
    
    articles_with_abstract = []
    pmids_no_abstract_recent = []  # 近90天无摘要
    pmids_no_abstract_old = []     # 超90天无摘要（丢弃）
    
    try:
        result_list = json_data.get("resultList", {}).get("result", [])
        logger.info(f"⚙️ Parsing {len(result_list)} articles from Europe PMC JSON...")
        
        for result in result_list:
            pmid = str(result.get("pmid", ""))
            abstract = result.get("abstractText", "")
            pub_date = result.get("firstPublicationDate", "")
            
            if not pmid:
                continue
            
            # 检查是否有摘要
            if abstract and abstract.strip():
                # 有摘要，提取完整元数据
                metadata = extract_europepmc_metadata(result)
                articles_with_abstract.append(metadata)
            else:
                # 无摘要，根据发表日期分类
                if EuropePMCClient.is_within_days(pub_date, 90):
                    pmids_no_abstract_recent.append(pmid)
                else:
                    pmids_no_abstract_old.append(pmid)
        
        logger.info(f"✅ Parsed: {len(articles_with_abstract)} with abstract, "
                   f"{len(pmids_no_abstract_recent)} no abstract (recent), "
                   f"{len(pmids_no_abstract_old)} no abstract (old, discarded)")
        
        return articles_with_abstract, pmids_no_abstract_recent, pmids_no_abstract_old
        
    except Exception as e:
        logger.error(f"❌ Failed to parse Europe PMC JSON: {e}", exc_info=True)
        return [], [], []


def extract_europepmc_metadata(result: Dict[str, Any]) -> Dict[str, Optional[str]]:
    """
    从 Europe PMC 单篇文献结果中提取元数据
    
    输出结构与 extract_article_metadata() 保持一致，
    确保后续处理流程无需修改。
    
    Args:
        result: Europe PMC API 返回的单篇文献 JSON
    
    Returns:
        标准化的元数据字典
    """
    import html
    from ..clients.EuropePMCClient import EuropePMCClient
    
    # 基本标识符
    pmid = str(result.get("pmid", ""))
    pmcid = result.get("pmcid", "")
    doi = result.get("doi", "")
    
    # 标题（需要清理 HTML）
    raw_title = result.get("title", "")
    title = EuropePMCClient.clean_title(raw_title)
    
    # 期刊信息
    journal_info = result.get("journalInfo", {})
    journal_obj = journal_info.get("journal", {})
    journal = journal_obj.get("title", "")
    issn = journal_obj.get("issn", "") or journal_obj.get("essn", "")
    journal_abbrev = journal_obj.get("isoabbreviation", "")
    
    # 作者信息
    author_list = result.get("authorList", {}).get("author", [])
    first_author = None
    corresponding_author = None
    first_affiliation = None
    all_authors = []
    
    if author_list:
        if isinstance(author_list, list) and len(author_list) > 0:
            # 第一作者
            first_author_data = author_list[0]
            fname = first_author_data.get("firstName", "") or ""
            lname = first_author_data.get("lastName", "") or ""
            first_author = f"{fname} {lname}".strip() or None
            
            # 第一作者单位
            affiliation_list = first_author_data.get("authorAffiliationDetailsList", {}).get("authorAffiliation", [])
            if affiliation_list:
                if isinstance(affiliation_list, list) and len(affiliation_list) > 0:
                    first_affiliation = affiliation_list[0].get("affiliation", "")
                elif isinstance(affiliation_list, dict):
                    first_affiliation = affiliation_list.get("affiliation", "")
            
            # 通讯作者（取最后一位）
            last_author_data = author_list[-1]
            fname = last_author_data.get("firstName", "") or ""
            lname = last_author_data.get("lastName", "") or ""
            corresponding_author = f"{fname} {lname}".strip() or None
            
            # 所有作者（RIS 导出用）
            for author in author_list:
                fname = author.get("firstName", "") or ""
                lname = author.get("lastName", "") or ""
                author_name = f"{lname}, {fname}".strip(", ") if lname else fname
                if author_name:
                    all_authors.append(author_name)
        
        elif isinstance(author_list, dict):
            # 单作者情况
            fname = author_list.get("firstName", "") or ""
            lname = author_list.get("lastName", "") or ""
            author_name = f"{fname} {lname}".strip() or None
            first_author = author_name
            corresponding_author = author_name
            
            if author_name:
                all_authors.append(f"{lname}, {fname}".strip(", ") if lname else fname)
            
            affiliation_list = author_list.get("authorAffiliationDetailsList", {}).get("authorAffiliation", [])
            if affiliation_list:
                if isinstance(affiliation_list, list) and len(affiliation_list) > 0:
                    first_affiliation = affiliation_list[0].get("affiliation", "")
    
    # 摘要（需要清理 HTML）
    raw_abstract = result.get("abstractText", "")
    abstract = EuropePMCClient.clean_title(raw_abstract) if raw_abstract else "No abstract available."
    
    # 发表日期
    pub_date = result.get("firstPublicationDate", "")  # 格式: YYYY-MM-DD
    
    # 卷、期、页码
    volume = journal_info.get("volume", "")
    issue = journal_info.get("issue", "")
    pages = result.get("pageInfo", "")
    
    # 语言
    language = result.get("language", "")
    
    # 关键词
    keywords = []
    keyword_list = result.get("keywordList", {}).get("keyword", [])
    if keyword_list:
        if isinstance(keyword_list, list):
            keywords = keyword_list
        else:
            keywords = [keyword_list]
    
    return {
        "pmid": pmid,
        "pmcid": pmcid,
        "title": title,
        "journal": journal,
        "issn": issn,
        "first_author": first_author,
        "corresponding_author": corresponding_author,
        "first_affiliation": first_affiliation,
        "abstract": abstract,
        "publication_date": pub_date,
        # RIS additional fields
        "doi": doi,
        "volume": volume,
        "issue": issue,
        "pages": pages,
        "journal_abbrev": journal_abbrev,
        "language": language,
        "keywords": keywords,
        "all_authors": all_authors,
        # 标记数据来源（用于调试）
        "_source": "europepmc"
    }

def _safe_get(data: Dict, path: List[str], default: Any = None) -> Any:
    """Safely get a value from a nested dictionary."""
    for key in path:
        if not isinstance(data, dict) or key not in data:
            return default
        data = data[key]
    return data

def extract_article_metadata(article_dict: Dict[str, Any]) -> Dict[str, Optional[str]]:
    """
    Extracts key metadata from a single article dictionary with robust title extraction.
    Also extracts additional fields for RIS export (DOI, Volume, Issue, Pages, Keywords, etc.)
    """
    pmid = _safe_get(article_dict, ["MedlineCitation", "PMID", "#text"])

    # --- PMCID and DOI Extraction ---
    pmcid = None
    doi = None
    article_id_list = _safe_get(article_dict, ["PubmedData", "ArticleIdList", "ArticleId"])
    if article_id_list:
        if isinstance(article_id_list, list):
            for article_id in article_id_list:
                if isinstance(article_id, dict):
                    if article_id.get("@IdType") == "pmc":
                        pmcid = article_id.get("#text")
                    elif article_id.get("@IdType") == "doi":
                        doi = article_id.get("#text")
        elif isinstance(article_id_list, dict):
            if article_id_list.get("@IdType") == "pmc":
                pmcid = article_id_list.get("#text")
            elif article_id_list.get("@IdType") == "doi":
                doi = article_id_list.get("#text")

    # --- Robust Title Extraction ---
    title_node = _safe_get(article_dict, ["MedlineCitation", "Article", "ArticleTitle"])
    title = None
    if isinstance(title_node, dict):
        title = title_node.get("#text")
    elif isinstance(title_node, str):
        title = title_node
    
    journal = _safe_get(article_dict, ["MedlineCitation", "Article", "Journal", "Title"])
    issn = _safe_get(article_dict, ["MedlineCitation", "Article", "Journal", "ISSN", "#text"])
    authors_data = _safe_get(article_dict, ["MedlineCitation", "Article", "AuthorList", "Author"])
    first_author = None
    corresponding_author = None
    first_affiliation = None

    if authors_data:
        # 处理多作者情况（list）
        if isinstance(authors_data, list) and len(authors_data) > 0:
            first_author_data = authors_data[0]
            if first_author_data:
                fname = _safe_get(first_author_data, ["ForeName"]) or ""
                lname = _safe_get(first_author_data, ["LastName"]) or ""
                first_author = f"{fname} {lname}".strip() or None
                affiliation_data = _safe_get(first_author_data, ["AffiliationInfo"])
                if affiliation_data:
                    if isinstance(affiliation_data, list) and len(affiliation_data) > 0:
                        first_affiliation = _safe_get(affiliation_data[0], ["Affiliation"])
                    else:
                        first_affiliation = _safe_get(affiliation_data, ["Affiliation"])
            last_author_data = authors_data[-1]
            if last_author_data:
                if _safe_get(last_author_data, ["CollectiveName"]):
                    corresponding_author = _safe_get(last_author_data, ["CollectiveName"])
                else:
                    fname = _safe_get(last_author_data, ["ForeName"]) or ""
                    lname = _safe_get(last_author_data, ["LastName"]) or ""
                    corresponding_author = f"{fname} {lname}".strip() or None

        # 处理单作者情况（dict）
        elif isinstance(authors_data, dict):
            # 提取唯一作者的信息
            fname = _safe_get(authors_data, ["ForeName"]) or ""
            lname = _safe_get(authors_data, ["LastName"]) or ""
            author_name = f"{fname} {lname}".strip() or None

            # 单作者既是第一作者也是通讯作者
            first_author = author_name
            corresponding_author = author_name

            # 提取单位信息
            affiliation_data = _safe_get(authors_data, ["AffiliationInfo"])
            if affiliation_data:
                if isinstance(affiliation_data, list) and len(affiliation_data) > 0:
                    first_affiliation = _safe_get(affiliation_data[0], ["Affiliation"])
                else:
                    first_affiliation = _safe_get(affiliation_data, ["Affiliation"])

    abstract_parts = []
    abstract_node = _safe_get(article_dict, ["MedlineCitation", "Article", "Abstract", "AbstractText"])
    if abstract_node:
        if isinstance(abstract_node, list):
            for part in abstract_node:
                label = part.get('@Label', '')
                text = part.get('#text', '')
                abstract_parts.append(f"[{label}] {text}" if label else text)
        elif isinstance(abstract_node, dict):
            label = abstract_node.get('@Label', '')
            text = abstract_node.get('#text', '')
            abstract_parts.append(f"[{label}] {text}" if label else text)
        else:
            abstract_parts.append(abstract_node)
    abstract = "\\n\\n".join(abstract_parts)

    # --- Publication Date ---
    journal_issue_node = _safe_get(article_dict, ["MedlineCitation", "Article", "Journal", "JournalIssue"])
    pub_date_node = _safe_get(journal_issue_node, ["PubDate"])
    year = _safe_get(pub_date_node, ["Year"])
    month = _safe_get(pub_date_node, ["Month"])
    day = _safe_get(pub_date_node, ["Day"])
    formatted_date = None
    if year:
        if month and day:
            month_map = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
            month_num = month_map.get(month, month)
            if len(str(month_num)) == 1: month_num = '0' + str(month_num)
            if len(str(day)) == 1: day = '0' + str(day)
            formatted_date = f"{year}-{month_num}-{day}"
        elif month:
            month_map = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
            month_num = month_map.get(month, month)
            if len(str(month_num)) == 1: month_num = '0' + str(month_num)
            formatted_date = f"{year}-{month_num}"
        else:
            formatted_date = str(year)

    # --- RIS Additional Fields ---
    # Volume and Issue
    volume = _safe_get(journal_issue_node, ["Volume"])
    issue = _safe_get(journal_issue_node, ["Issue"])

    # Pages
    pages = _safe_get(article_dict, ["MedlineCitation", "Article", "Pagination", "MedlinePgn"])

    # Journal Abbreviation
    journal_abbrev = _safe_get(article_dict, ["MedlineCitation", "Article", "Journal", "ISOAbbreviation"])

    # Language
    language = _safe_get(article_dict, ["MedlineCitation", "Article", "Language"])

    # Keywords
    keywords = []
    keyword_list = _safe_get(article_dict, ["MedlineCitation", "KeywordList", "Keyword"])
    if keyword_list:
        if isinstance(keyword_list, list):
            for kw in keyword_list:
                if isinstance(kw, dict):
                    keywords.append(kw.get("#text", ""))
                elif isinstance(kw, str):
                    keywords.append(kw)
        elif isinstance(keyword_list, dict):
            keywords.append(keyword_list.get("#text", ""))
        elif isinstance(keyword_list, str):
            keywords.append(keyword_list)

    # All Authors (for RIS export)
    all_authors = []
    if authors_data:
        if isinstance(authors_data, list):
            for author in authors_data:
                if isinstance(author, dict):
                    if _safe_get(author, ["CollectiveName"]):
                        all_authors.append(_safe_get(author, ["CollectiveName"]))
                    else:
                        fname = _safe_get(author, ["ForeName"]) or ""
                        lname = _safe_get(author, ["LastName"]) or ""
                        author_name = f"{lname}, {fname}".strip(", ") if lname else fname
                        if author_name:
                            all_authors.append(author_name)
        elif isinstance(authors_data, dict):
            if _safe_get(authors_data, ["CollectiveName"]):
                all_authors.append(_safe_get(authors_data, ["CollectiveName"]))
            else:
                fname = _safe_get(authors_data, ["ForeName"]) or ""
                lname = _safe_get(authors_data, ["LastName"]) or ""
                author_name = f"{lname}, {fname}".strip(", ") if lname else fname
                if author_name:
                    all_authors.append(author_name)

    return {
        "pmid": pmid, "pmcid": pmcid, "title": title, "journal": journal, "issn": issn,
        "first_author": first_author, "corresponding_author": corresponding_author,
        "first_affiliation": first_affiliation, "abstract": abstract or "No abstract available.",
        "publication_date": formatted_date,
        # RIS additional fields
        "doi": doi,
        "volume": volume,
        "issue": issue,
        "pages": pages,
        "journal_abbrev": journal_abbrev,
        "language": language,
        "keywords": keywords,
        "all_authors": all_authors,
    }

def _load_journal_ref_to_dict(reference_file: str) -> Dict[str, Dict[str, str]]:
    """Loads the journal reference CSV into a dictionary keyed by normalized ISSN."""
    journal_dict = {}
    try:
        # 尝试UTF-8编码，如果失败则尝试GBK编码
        try:
            with open(reference_file, 'r', encoding='utf-8') as f:
                content = f.read()
            encoding = 'utf-8'
        except UnicodeDecodeError:
            with open(reference_file, 'r', encoding='gbk') as f:
                content = f.read()
            encoding = 'gbk'

        # 使用检测到的编码重新打开文件
        with open(reference_file, 'r', encoding=encoding) as f:
            reader = csv.reader(f)
            next(reader)  # Skip header
            for row in reader:
                if len(row) < 11: continue

                # 新文件列名映射：Journal、缩写、ISSN、EISSN、Catalog、Publisher、CAS-zone、JCR-zone、2024-IF、5-year-IF、JIF Rank
                journal_info = {
                    '期刊名称_ref': row[0].strip(),      # Journal
                    '中科院分区': row[6].strip(),          # CAS-zone (中科院分区)
                    'JCR分区': row[7].strip(),           # JCR-zone
                    '最新IF': row[8].strip(),            # 2024-IF
                    '5年IF': row[9].strip(),             # 5-year-IF
                    '排名': row[10].strip()              # JIF Rank
                }

                print_issn = row[2].strip().replace('-', '')  # ISSN
                e_issn = row[3].strip().replace('-', '')      # EISSN

                if print_issn:
                    journal_dict[print_issn] = journal_info
                if e_issn:
                    journal_dict[e_issn] = journal_info
        logger.info(f"📚 Successfully loaded {len(journal_dict)} journal entries into dictionary.")
    except Exception as e:
        logger.error(f"❌ Failed to load reference file into dictionary: {e}", exc_info=True)
    return journal_dict


def enrich_articles_with_journal_info(
    screened_articles: List[Dict[str, Any]],
    reference_file: str,
    logger: logging.Logger
) -> List[Dict[str, Any]]:
    """
    Enriches articles with journal info from CSV using the logic that powers Excel generation.
    Returns a NEW list of enriched articles.
    """
    if not screened_articles:
        return []

    journal_ref_dict = _load_journal_ref_to_dict(reference_file)
    # logger.info(f"🔍 DEBUG: Loaded journal ref dict with {len(journal_ref_dict)} entries")
    
    enriched_articles = []

    for article in screened_articles:
        enriched_article = article.copy()
        issn = article.get("issn", "")
        normalized_issn = issn.replace('-', '').strip() if issn else ""

        journal_info = journal_ref_dict.get(normalized_issn, {})
        
        if journal_info:
            logger.info(f"✅ DEBUG: Found journal info for ISSN {normalized_issn}: {journal_info.get('最新IF')}")
        else:
            if normalized_issn:
                logger.info(f"⚠️ DEBUG: No journal info found for ISSN {normalized_issn} (Original: {issn})")

        # Use keys matching _load_journal_ref_to_dict
        enriched_article['journal'] = journal_info.get('期刊名称_ref', article.get('journal'))
        enriched_article['cas_zone'] = journal_info.get('中科院分区', enriched_article.get('cas_zone', ''))
        enriched_article['jcr_zone'] = journal_info.get('JCR分区', enriched_article.get('jcr_zone', ''))
        enriched_article['latest_if'] = journal_info.get('最新IF', enriched_article.get('latest_if', ''))
        enriched_article['five_year_if'] = journal_info.get('5年IF', enriched_article.get('five_year_if', ''))
        enriched_article['ranking'] = journal_info.get('排名', enriched_article.get('ranking', ''))

        # Add PubMed Link
        pmid = enriched_article.get('pmid')
        if pmid:
            enriched_article['pubmed_link'] = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
        else:
            enriched_article['pubmed_link'] = ""

        # Add PMC Link logic (simplified reuse)
        existing_pmc_link = enriched_article.get('pmc_link') or article.get('pmc_link')
        if existing_pmc_link and str(existing_pmc_link).startswith('http'):
            enriched_article['pmc_link'] = existing_pmc_link
        else:
            pmcid = enriched_article.get('pmcid')
            if pmcid:
                enriched_article['pmc_link'] = f"https://pmc.ncbi.nlm.nih.gov/articles/{pmcid}/"
            elif enriched_article.get('oa_link') or article.get('oa_link'):
                enriched_article['pmc_link'] = enriched_article.get('oa_link') or article.get('oa_link')
            else:
                enriched_article['pmc_link'] = ""

        enriched_articles.append(enriched_article)

    return enriched_articles

def generate_formatted_excel(screened_articles: List[Dict[str, Any]], reference_file: str, output_path: str, language_config=None):
    """
    Generates a fully formatted Excel file, enriching articles with journal data.
    Uses English key names internally, maps to display names via language_config.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    if not screened_articles:
        logger.warning("⚠️ No articles to process. Excel file will not be generated.")
        return

    # Use the shared enrichment function
    enriched_articles = enrich_articles_with_journal_info(screened_articles, reference_file, logger)


    # 📊 OA 统计：计算所有有 OA 链接的文章
    total_articles = len(enriched_articles)
    pmc_count = 0
    unpaywall_count = 0
    no_oa_count = 0
    
    for article in enriched_articles:
        pmc_link = article.get('pmc_link', '')
        if pmc_link:
            if 'pmc.ncbi.nlm.nih.gov' in pmc_link:
                pmc_count += 1
            else:
                unpaywall_count += 1  # 其他 OA 链接（Unpaywall 或其他来源）
        else:
            no_oa_count += 1
    
    total_oa = pmc_count + unpaywall_count
    oa_rate = (total_oa / total_articles * 100) if total_articles > 0 else 0
    
    logger.info(f"📊 OA Statistics: {total_oa}/{total_articles} articles ({oa_rate:.1f}% OA rate)")
    logger.info(f"   • PMC links: {pmc_count} | Other OA: {unpaywall_count} | No OA: {no_oa_count}")

    final_df = pd.DataFrame(enriched_articles)

    # 英文键名列表（内部使用）
    english_columns = [
        "index", "journal", "title", "publication_date",
        "research_objective", "study_type", "research_method", "study_population",
        "main_results", "conclusions", "highlights",
        "first_author", "corresponding_author", "first_author_affiliation",
        "issn", "cas_zone", "jcr_zone", "latest_if", "five_year_if", "ranking",
        "pmid", "pubmed_link", "pmc_link"
    ]

    # 🌍 根据语言配置设置显示列名
    # 英文键名 → 中文键名的映射（language_config.json 中使用中文键名）
    column_key_mapping = {
        "index": "序号",
        "journal": "期刊名称",
        "title": "文章标题",
        "publication_date": "发表时间",
        "research_objective": "研究目的",
        "study_type": "研究类型",
        "research_method": "研究方法",
        "study_population": "研究对象",
        "main_results": "主要研究结果",
        "conclusions": "研究结论与意义",
        "highlights": "研究亮点或创新点",
        "first_author": "第一作者",
        "corresponding_author": "通讯作者",
        "first_author_affiliation": "第一作者单位",
        "issn": "ISSN",
        "cas_zone": "中科院分区",
        "jcr_zone": "JCR分区",
        "latest_if": "最新IF",
        "five_year_if": "5年IF",
        "ranking": "排名",
        "pmid": "PMID",
        "pubmed_link": "PubMed链接",
        "pmc_link": "PMC链接"
    }
    
    # 默认英文列名
    default_columns = {
        "index": "No.",
        "journal": "Journal",
        "title": "Title",
        "publication_date": "Publication Date",
        "research_objective": "Research Objective",
        "study_type": "Study Type",
        "research_method": "Research Method",
        "study_population": "Study Population",
        "main_results": "Main Results",
        "conclusions": "Conclusions",
        "highlights": "Highlights",
        "first_author": "First Author",
        "corresponding_author": "Corresponding Author",
        "first_author_affiliation": "First Author Affiliation",
        "issn": "ISSN",
        "cas_zone": "CAS Zone",
        "jcr_zone": "JCR Zone",
        "latest_if": "Latest IF",
        "five_year_if": "5-Year IF",
        "ranking": "Ranking",
        "pmid": "PMID",
        "pubmed_link": "PubMed Link",
        "pmc_link": "PMC Link"
    }
    
    if language_config and 'fields' in language_config:
        fields = language_config['fields']
        final_columns = []
        for eng_col in english_columns:
            # 通过映射表获取中文键名，再从 fields 中获取翻译后的列名
            zh_key = column_key_mapping.get(eng_col, eng_col)
            translated_col = fields.get(zh_key, default_columns.get(eng_col, eng_col))
            final_columns.append(translated_col)
    else:
        # 默认英文列名
        final_columns = [
            "No.", "Journal", "Title", "Publication Date",
            "Research Objective", "Study Type", "Research Method", "Study Population",
            "Main Results", "Conclusions", "Highlights",
            "First Author", "Corresponding Author", "First Author Affiliation",
            "ISSN", "CAS Zone", "JCR Zone", "Latest IF", "5-Year IF", "Ranking",
            "PMID", "PubMed Link", "PMC Link"
        ]

    # 🌍 按评分分组（3分、2分、1分）
    score_groups = {}
    for article in enriched_articles:
        score = article.get('score', '')
        # 提取纯数字评分
        if '3' in str(score) or score == '3':
            score_key = '3'
        elif '2' in str(score) or score == '2':
            score_key = '2'
        elif '1' in str(score) or score == '1':
            score_key = '1'
        else:
            score_key = '1'  # 默认归为1分

        if score_key not in score_groups:
            score_groups[score_key] = []
        score_groups[score_key].append(article)

    # 🌍 获取评分等级的sheet名称
    scoring_levels = language_config.get('scoring_levels', {
        "3": "3分 (高度相关)",
        "2": "2分 (中度相关)",
        "1": "1分 (轻度相关)"
    }) if language_config else {
        "3": "3分 (高度相关)",
        "2": "2分 (中度相关)",
        "1": "1分 (轻度相关)"
    }

    try:
        logger.info(f"📊 Writing formatted data to Excel file: {output_path}")
        wb = Workbook()

        # 删除默认的sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        cell_alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        hyperlink_font = Font(color="0563C1", underline="single")

        # 按评分从高到低创建sheet（3分 → 2分 → 1分）
        for score_key in ['3', '2', '1']:
            if score_key not in score_groups or not score_groups[score_key]:
                continue  # 跳过没有文章的评分等级

            # 获取sheet名称（使用对应语言）
            sheet_name = scoring_levels.get(score_key, f"{score_key}分")
            # Excel sheet名称最长31个字符
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            ws = wb.create_sheet(title=sheet_name)

            # 创建该评分等级的DataFrame
            score_df = pd.DataFrame(score_groups[score_key])

            # 重新排列列（使用英文键名）
            score_df = score_df.reindex(columns=english_columns)

            # 重新编号
            score_df['index'] = range(1, len(score_df) + 1)

            # 🌍 如果有语言配置，重命名列名为显示名
            if language_config and 'fields' in language_config:
                fields = language_config['fields']
                column_mapping = {}
                for eng_col in english_columns:
                    # 通过映射表获取中文键名，再从 fields 中获取翻译后的列名
                    zh_key = column_key_mapping.get(eng_col, eng_col)
                    translated_col = fields.get(zh_key, default_columns.get(eng_col, eng_col))
                    if translated_col != eng_col:
                        column_mapping[eng_col] = translated_col

                if column_mapping:
                    score_df = score_df.rename(columns=column_mapping)

            # 写入数据（清理非法字符）
            for r_idx, row in enumerate(dataframe_to_rows(score_df, index=False, header=True), 1):
                # 清理每个单元格中的非法字符
                cleaned_row = [clean_illegal_characters(cell) for cell in row]
                ws.append(cleaned_row)
                for cell in ws[r_idx]:
                    cell.alignment = cell_alignment
                    cell.border = thin_border

                if r_idx % 2 != 0:  # Apply alternating row color to ODD rows (1, 3, 5...)
                    for cell in ws[r_idx]:
                        cell.fill = green_fill

                # Add hyperlinks for both PubMed and PMC links
                if r_idx > 1:
                    # PubMed链接 (倒数第二列)
                    pubmed_link_cell = ws.cell(row=r_idx, column=len(final_columns) - 1)
                    if pubmed_link_cell.value and pubmed_link_cell.value.startswith('http'):
                        pubmed_link_cell.hyperlink = pubmed_link_cell.value
                        pubmed_link_cell.font = hyperlink_font

                    # PMC链接 (最后一列)
                    pmc_link_cell = ws.cell(row=r_idx, column=len(final_columns))
                    if pmc_link_cell.value and pmc_link_cell.value.startswith('http'):
                        pmc_link_cell.hyperlink = pmc_link_cell.value
                        pmc_link_cell.font = hyperlink_font

            # Set column widths
            # 顺序: index → journal → title → publication_date → research_objective → study_type →
            #       research_method → study_population → main_results → conclusions → highlights →
            #       first_author → corresponding_author → first_author_affiliation →
            #       issn → cas_zone → jcr_zone → latest_if → five_year_if → ranking → pmid → pubmed_link → pmc_link
            column_widths = {
                'A': 5,   # index
                'B': 20,  # journal
                'C': 40,  # title
                'D': 12,  # publication_date
                'E': 50,  # research_objective
                'F': 20,  # study_type
                'G': 50,  # research_method
                'H': 30,  # study_population
                'I': 50,  # main_results
                'J': 50,  # conclusions
                'K': 50,  # highlights
                'L': 15,  # first_author
                'M': 15,  # corresponding_author
                'N': 30,  # first_author_affiliation
                'O': 12,  # issn
                'P': 8,   # cas_zone
                'Q': 8,   # jcr_zone
                'R': 8,   # latest_if
                'S': 8,   # five_year_if
                'T': 10,  # ranking
                'U': 10,  # pmid
                'V': 25,  # pubmed_link
                'W': 25   # pmc_link
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            ws.freeze_panes = 'B2'  # 冻结首行和首列（序号列）
            ws.auto_filter.ref = ws.dimensions

            logger.info(f"✅ Created sheet '{sheet_name}' with {len(score_df)} articles")

        wb.save(output_path)
        total_articles = sum(len(articles) for articles in score_groups.values())
        logger.info(f"✅ Successfully saved {total_articles} articles to {output_path} ({len(score_groups)} sheets)")

    except Exception as e:
        logger.error(f"❌ Failed to write formatted Excel file: {e}", exc_info=True)


def generate_ris_file(screened_articles: List[Dict], output_path: str) -> None:
    """
    Generate RIS format file from screened articles.

    RIS (Research Information Systems) format is a standardized tag format
    for exchanging bibliographic data, supported by EndNote, Zotero, Mendeley, etc.

    Args:
        screened_articles: List of article dictionaries with metadata (using English key names)
        output_path: Path to save the RIS file
    """
    if not screened_articles:
        logger.warning("⚠️ No articles to export to RIS format.")
        return

    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            for article in screened_articles:
                # TY - Type of reference (Journal Article)
                f.write("TY  - JOUR\n")

                # TI - Title
                title = article.get("title", "")
                if title:
                    f.write(f"TI  - {title}\n")

                # AU - Authors (one per line)
                all_authors = article.get("all_authors", [])
                if all_authors:
                    for author in all_authors:
                        if author:
                            f.write(f"AU  - {author}\n")
                else:
                    # Fallback to first author and corresponding author
                    first_author = article.get("first_author", "")
                    if first_author:
                        f.write(f"AU  - {first_author}\n")
                    corresponding = article.get("corresponding_author", "")
                    if corresponding and corresponding != first_author:
                        f.write(f"AU  - {corresponding}\n")

                # JO - Journal full name
                journal = article.get("journal", "")
                if journal:
                    f.write(f"JO  - {journal}\n")

                # JA - Journal abbreviation
                journal_abbrev = article.get("journal_abbrev", "")
                if journal_abbrev:
                    f.write(f"JA  - {journal_abbrev}\n")

                # PY - Publication year
                pub_date = article.get("publication_date", "")
                if pub_date:
                    year = str(pub_date)[:4] if pub_date else ""
                    if year:
                        f.write(f"PY  - {year}\n")
                    # DA - Full date
                    f.write(f"DA  - {pub_date}\n")

                # VL - Volume
                volume = article.get("volume", "")
                if volume:
                    f.write(f"VL  - {volume}\n")

                # IS - Issue
                issue = article.get("issue", "")
                if issue:
                    f.write(f"IS  - {issue}\n")

                # SP/EP - Start/End Page
                pages = article.get("pages", "")
                if pages:
                    if "-" in str(pages):
                        parts = str(pages).split("-")
                        f.write(f"SP  - {parts[0].strip()}\n")
                        if len(parts) > 1:
                            f.write(f"EP  - {parts[1].strip()}\n")
                    else:
                        f.write(f"SP  - {pages}\n")

                # AB - Abstract
                abstract = article.get("abstract", "")
                if abstract and abstract != "No abstract available.":
                    # Replace escaped newlines with actual spaces for RIS format
                    clean_abstract = abstract.replace("\\n", " ").replace("\n", " ")
                    f.write(f"AB  - {clean_abstract}\n")

                # KW - Keywords
                keywords = article.get("keywords", [])
                if keywords:
                    for kw in keywords:
                        if kw:
                            f.write(f"KW  - {kw}\n")

                # SN - ISSN
                issn = article.get("issn", "")
                if issn:
                    f.write(f"SN  - {issn}\n")

                # LA - Language
                language = article.get("language", "")
                if language:
                    f.write(f"LA  - {language}\n")

                # AN - Accession Number (PMID)
                pmid = article.get("pmid", "")
                if pmid:
                    f.write(f"AN  - {pmid}\n")

                # DO - DOI
                doi = article.get("doi", "")
                if doi:
                    f.write(f"DO  - {doi}\n")

                # UR - URL (PubMed link)
                if pmid:
                    f.write(f"UR  - https://pubmed.ncbi.nlm.nih.gov/{pmid}/\n")

                # N1 - Notes (Relevance Score)
                score = article.get("score", "")
                if score:
                    # Extract numeric score from formatted string (e.g., "3分 (高度相关)" -> "3")
                    score_num = ""
                    if score and len(score) > 0:
                        score_num = score[0] if score[0].isdigit() else ""

                    if score_num == "3":
                        f.write("N1  - Score 3/3: Highly Relevant\n")
                    elif score_num == "2":
                        f.write("N1  - Score 2/3: Moderately Relevant\n")
                    elif score_num == "1":
                        f.write("N1  - Score 1/3: Marginally Relevant\n")
                    else:
                        f.write(f"N1  - Score: {score}\n")

                # ER - End of Reference
                f.write("ER  - \n\n")

        logger.info(f"✅ Successfully exported {len(screened_articles)} articles to RIS: {output_path}")

    except Exception as e:
        logger.error(f"❌ Failed to generate RIS file: {e}", exc_info=True)


# ============================================================
# Unpaywall API 相关函数
# ============================================================

def get_oa_link_from_unpaywall(doi: str, exclude_bronze: bool = True) -> Optional[str]:
    """
    通过 Unpaywall API 获取文章的 OA 链接

    Args:
        doi: 文章的 DOI
        exclude_bronze: 是否排除 bronze 类型（可访问但无开放许可证）

    Returns:
        OA 链接（如果有），否则返回 None
    """
    if not doi:
        return None

    MAX_RETRIES = 2
    import time
    
    for attempt in range(MAX_RETRIES):
        try:
            # 🔄 使用轮换的 email
            email = UnpaywallEmailManager.get_next_email()
            url = f"https://api.unpaywall.org/v2/{doi}?email={email}"
            response = requests.get(url, timeout=5)

            if response.status_code == 200:
                data = response.json()

                if data.get("is_oa"):
                    oa_status = data.get("oa_status")

                    # 排除 bronze（可访问但非真正开放，可能随时变付费）
                    if exclude_bronze and oa_status == "bronze":
                        logger.debug(f"🚫 Skipping bronze OA for DOI: {doi}")
                        return None

                    best_location = data.get("best_oa_location", {})
                    if best_location:
                        # 优先返回 PDF 直链，否则返回普通链接
                        oa_url = best_location.get("url_for_pdf") or best_location.get("url")
                        if oa_url:
                            return oa_url
                
                # 成功请求但没有 OA 链接，直接返回 None，不重试
                return None

            # 404 等明确错误不重试
            elif response.status_code == 404:
                return None

        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
            if attempt < MAX_RETRIES - 1:
                logger.debug(f"⚠️ Unpaywall connection error/timeout for DOI {doi} (Attempt {attempt + 1}/{MAX_RETRIES}). Retrying...")
                time.sleep(1)
            else:
                logger.warning(f"❌ Unpaywall failed for DOI {doi} after {MAX_RETRIES} attempts: {e}")
                return None
        except Exception as e:
            logger.debug(f"⚠️ Unpaywall error for DOI {doi}: {e}")
            return None
    
    return None